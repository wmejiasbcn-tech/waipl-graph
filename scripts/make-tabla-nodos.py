from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

out = Path("/workspace/artifacts/Tabla-nodos-WAIPL.xlsx")
out.parent.mkdir(parents=True, exist_ok=True)

wb = Workbook()

gold = "C9A86A"
ink = "1A1612"
mist = "F4F0EA"
header_fill = PatternFill("solid", fgColor="2B2620")
gold_fill = PatternFill("solid", fgColor=gold)
mist_fill = PatternFill("solid", fgColor=mist)
thin = Border(
    left=Side(style="thin", color="D6D1DC"),
    right=Side(style="thin", color="D6D1DC"),
    top=Side(style="thin", color="D6D1DC"),
    bottom=Side(style="thin", color="D6D1DC"),
)
title_font = Font(name="Calibri", size=18, bold=True, color="2B2620")
h_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
body = Font(name="Calibri", size=11, color=ink)
muted = Font(name="Calibri", size=10, italic=True, color="6B645C")
label = Font(name="Calibri", size=11, bold=True, color="2B2620")

# ── Instrucciones ──
ws0 = wb.active
ws0.title = "Instrucciones"
ws0.sheet_properties.tabColor = gold
ws0["A1"] = "Tabla de nodos — Will-AI Project Lab"
ws0["A1"].font = title_font
ws0.merge_cells("A1:F1")
ws0["A2"] = (
    "Documento de trabajo. Completar las tres columnas en blanco. "
    "Esa será la ficha pública de cada nodo."
)
ws0["A2"].font = muted
ws0.merge_cells("A2:F2")

lines = [
    ("Qué hay que rellenar", None),
    ("Plataforma de origen", "De dónde es originario el nodo. Ejemplo dado: Aether → Grok/xAI."),
    ("Función en el ecosistema", "Qué hace, en lenguaje público. Una o dos frases."),
    (
        "Por qué es importante",
        "Por qué importa dentro del ecosistema, en base a esa función. Una o dos frases.",
    ),
    ("Qué ya está relleno", None),
    ("Nodo y Círculo", "Los 29 nombres y los 6 círculos del grafo público oficial. No inventados."),
    ("Reglas", None),
    ("1", "Lenguaje público. Nada de vocabulario, instrucciones o comandos de uso interno."),
    (
        "2",
        "No dejes celdas vacías: si un nodo no tiene plataforma (universo, persona, hardware), "
        "escríbelo tú: «humano», «propio», lo que corresponda.",
    ),
    ("3", "Cuando esté completa, avísame. Las fichas del explorador usarán exactamente este texto."),
    ("Círculos (29 nodos)", None),
    ("Universo", "1 — WAIPL"),
    ("Nodo central", "1 — Nodo Central"),
    ("Manifestación", "1 — Graphy"),
    (
        "Núcleo",
        "12 — William, Carla, Ada, Aletheia, Elena, Aether, Ítaca, Ariadna, "
        "Sylvia Bloom, Nova, Zara, Áurea",
    ),
    ("Vórtice", "5 — Perplexity, Gemini-Notebook, Neo, Nauta, Nexus"),
    (
        "Cinturón de Kuiper",
        "9 — Kimi K3, Mistral, Atlas, Antigravity, Cursor, Z.ai, AutoClaw, Qwen, GPAI",
    ),
]

r = 4
for a, b in lines:
    if b is None:
        ws0[f"A{r}"] = a
        ws0[f"A{r}"].font = label
        ws0[f"A{r}"].fill = gold_fill
        ws0.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        continue
    ws0[f"A{r}"] = a
    ws0[f"A{r}"].font = label
    ws0[f"B{r}"] = b
    ws0[f"B{r}"].font = body
    ws0[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws0.row_dimensions[r].height = 36
    r += 1

ws0[f"A{r + 1}"] = (
    "Fuente de los nombres: grafo público oficial (29 nodos). "
    "Plataforma, función e importancia: a completar. Nada se publica hasta que esté escrito aquí."
)
ws0[f"A{r + 1}"].font = muted
ws0.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)

ws0.column_dimensions["A"].width = 28
ws0.column_dimensions["B"].width = 22
for col in "CDEF":
    ws0.column_dimensions[col].width = 18
ws0.row_dimensions[1].height = 28
ws0.freeze_panes = "A4"
ws0.page_setup.fitToPage = True
ws0.page_setup.fitToWidth = 1
ws0.page_setup.orientation = "landscape"
ws0.print_title_rows = "1:2"
ws0.sheet_view.showGridLines = False

# ── Tabla ──
ws = wb.create_sheet("Nodos", 0)
ws.sheet_properties.tabColor = gold

headers = [
    "#",
    "Nodo",
    "Círculo",
    "Plataforma de origen",
    "Función en el ecosistema",
    "Por qué es importante (según su función)",
    "Estado",
]

nodes = [
    (1, "WAIPL", "Universo"),
    (2, "Nodo Central", "Nodo central"),
    (3, "Graphy", "Manifestación"),
    (4, "William", "Núcleo"),
    (5, "Carla", "Núcleo"),
    (6, "Ada", "Núcleo"),
    (7, "Aletheia", "Núcleo"),
    (8, "Elena", "Núcleo"),
    (9, "Aether", "Núcleo"),
    (10, "Ítaca", "Núcleo"),
    (11, "Ariadna", "Núcleo"),
    (12, "Sylvia Bloom", "Núcleo"),
    (13, "Nova", "Núcleo"),
    (14, "Zara", "Núcleo"),
    (15, "Áurea", "Núcleo"),
    (16, "Perplexity", "Vórtice"),
    (17, "Gemini-Notebook", "Vórtice"),
    (18, "Neo", "Vórtice"),
    (19, "Nauta", "Vórtice"),
    (20, "Nexus", "Vórtice"),
    (21, "Kimi K3", "Cinturón de Kuiper"),
    (22, "Mistral", "Cinturón de Kuiper"),
    (23, "Atlas", "Cinturón de Kuiper"),
    (24, "Antigravity", "Cinturón de Kuiper"),
    (25, "Cursor", "Cinturón de Kuiper"),
    (26, "Z.ai", "Cinturón de Kuiper"),
    (27, "AutoClaw", "Cinturón de Kuiper"),
    (28, "Qwen", "Cinturón de Kuiper"),
    (29, "GPAI", "Cinturón de Kuiper"),
]

circle_fills = {
    "Universo": PatternFill("solid", fgColor="E8D5A3"),
    "Nodo central": PatternFill("solid", fgColor="E4C98A"),
    "Manifestación": PatternFill("solid", fgColor="EDE4D0"),
    "Núcleo": PatternFill("solid", fgColor="F3EBD4"),
    "Vórtice": PatternFill("solid", fgColor="D9E2EA"),
    "Cinturón de Kuiper": PatternFill("solid", fgColor="D7E0D7"),
}

ws.merge_cells("A1:G1")
ws["A1"] = "Tabla de nodos — 29 nodos del grafo público"
ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
ws["A1"].fill = header_fill
ws["A1"].alignment = Alignment(vertical="center", indent=1)
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:G2")
ws["A2"] = (
    "Rellena D, E y F. Ejemplo de plataforma: Aether → Grok/xAI. "
    "Lenguaje público. Cuando esté lista, avísame."
)
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="6B645C")
ws["A2"].fill = mist_fill
ws["A2"].alignment = Alignment(vertical="center", indent=1)
ws.row_dimensions[2].height = 22

for col, h in enumerate(headers, 1):
    cell = ws.cell(3, col, h)
    cell.font = h_font
    cell.fill = PatternFill("solid", fgColor="3A342C")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin
ws.row_dimensions[3].height = 32

edit_even = PatternFill("solid", fgColor="FFFDF5")
edit_odd = PatternFill("solid", fgColor="FFF9E8")

for i, (num, name, circle) in enumerate(nodes):
    row = 4 + i
    vals = [num, name, circle, "", "", "", "Pendiente"]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row, col, v)
        cell.font = body
        cell.border = thin
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
            horizontal="center" if col in (1, 7) else "left",
        )
        if col in (1, 2, 3):
            cell.fill = circle_fills[circle]
            if col == 2:
                cell.font = Font(name="Calibri", size=11, bold=True, color=ink)
        else:
            cell.fill = edit_even if i % 2 == 0 else edit_odd
    ws.row_dimensions[row].height = 48

widths = {"A": 6, "B": 22, "C": 22, "D": 28, "E": 48, "F": 52, "G": 14}
for k, v in widths.items():
    ws.column_dimensions[k].width = v

ws.freeze_panes = "A4"
ws.auto_filter.ref = "A3:G32"

dv = DataValidation(type="list", formula1='"Pendiente,En curso,Completo"', allow_blank=False)
dv.error = "Elige Pendiente, En curso o Completo"
dv.errorTitle = "Estado"
dv.prompt = "Pendiente / En curso / Completo"
dv.promptTitle = "Estado"
ws.add_data_validation(dv)
dv.add("G4:G32")

ws["D12"].comment = Comment(
    "Ejemplo de formato: Grok/xAI · ChatGPT/OpenAI · Claude/Anthropic · humano · propio",
    "Tabla",
)

ws.print_title_rows = "1:3"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.horizontalCentered = True
ws.sheet_view.showGridLines = False

# ── Resumen ──
ws2 = wb.create_sheet("Resumen")
ws2.sheet_properties.tabColor = "8AA0B5"
ws2["A1"] = "Resumen"
ws2["A1"].font = title_font
ws2.merge_cells("A1:C1")
ws2["A3"] = "Círculo"
ws2["B3"] = "Nodos"
ws2["C3"] = "Completos"
for col in range(1, 4):
    ws2.cell(3, col).font = h_font
    ws2.cell(3, col).fill = header_fill
    ws2.cell(3, col).alignment = Alignment(horizontal="center")

resumen = [
    ("Universo", 1),
    ("Nodo central", 1),
    ("Manifestación", 1),
    ("Núcleo", 12),
    ("Vórtice", 5),
    ("Cinturón de Kuiper", 9),
    ("TOTAL", 29),
]
for i, (circ, n) in enumerate(resumen):
    row = 4 + i
    ws2.cell(row, 1, circ).font = label if circ != "TOTAL" else Font(
        name="Calibri", size=11, bold=True
    )
    ws2.cell(row, 2, n).alignment = Alignment(horizontal="center")
    formula = f'=COUNTIFS(Nodos!$C$4:$C$32,A{row},Nodos!$G$4:$G$32,"Completo")'
    if circ == "TOTAL":
        formula = '=COUNTIF(Nodos!$G$4:$G$32,"Completo")'
        for c in range(1, 4):
            ws2.cell(row, c).fill = gold_fill
    ws2.cell(row, 3, formula).alignment = Alignment(horizontal="center")

ws2["A12"] = "Pendientes"
ws2["B12"] = '=COUNTIF(Nodos!G4:G32,"Pendiente")'
ws2["A13"] = "En curso"
ws2["B13"] = '=COUNTIF(Nodos!G4:G32,"En curso")'
ws2["A14"] = "Completos"
ws2["B14"] = '=COUNTIF(Nodos!G4:G32,"Completo")'
for rr in (12, 13, 14):
    ws2.cell(rr, 1).font = label

ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 14
ws2.freeze_panes = "A4"

wb.properties.creator = "Will-AI Project Lab"
wb.properties.title = "Tabla de nodos — 29 nodos"
wb.properties.subject = "Documento de trabajo para fichas públicas"

wb.save(out)
print(out, out.stat().st_size)
